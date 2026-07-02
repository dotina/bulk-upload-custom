import openpyxl
import pytest

from convert_espn_ct import convert_workbook


@pytest.fixture
def espn_workbook_path(tmp_path):
    path = tmp_path / "CT chargers.xlsx"
    workbook = openpyxl.Workbook()
    site_sheet = workbook.active
    site_sheet.title = "chargers (3)"
    site_sheet.append(["name", "address", "latitude", "longitude"])
    site_sheet.append(
        ["ESPN / ESPN 11-2", "1 Espn Plaza , Bristol, Connecticut, United States, 06010", 41.649, -72.900]
    )

    charger_sheet = workbook.create_sheet("CT chargers")
    charger_sheet.append(
        [
            "EVSE ID", "Display Name", "MAC Address", "EVSE Latitude", "EVSE Longitude",
            "Port 1: Voltage (V)", "Port 1: Current (A)", "Port 1: Connector Type",
            "Port 2: Voltage (V)", "Port 2: Current (A)", "Port 2: Connector Type",
            "EVSE Usage Category", "Serial Number",
        ]
    )
    charger_sheet.append(
        [
            220754, "ESPN / ESPN 11-2", "0024:B100:0002:872C", 41.650, -72.904,
            "240V", "30A", "J1772", "240V", "30A", "J1772",
            "Commercial with unrestricted access", 172841008201,
        ]
    )
    workbook.save(path)
    return str(path)


def test_convert_workbook_maps_site_name(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["site_name"] == "ESPN / ESPN 11-2"


def test_convert_workbook_splits_address_into_city(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["site_city"] == "Bristol"


def test_convert_workbook_strips_voltage_unit_suffix(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["port1_voltage_v"] == "240"


def test_convert_workbook_uses_default_contact_email(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["site_contact_email"] == "ops@example.com"


def test_convert_workbook_hardcodes_charger_type_level_2(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["charger_type"] == "Level 2"


def test_convert_workbook_maps_serial_number(espn_workbook_path):
    rows = convert_workbook(espn_workbook_path, default_contact_email="ops@example.com")
    assert rows[0]["serial_number"] == 172841008201


def test_convert_workbook_missing_matching_site_raises(tmp_path):
    path = tmp_path / "CT chargers.xlsx"
    workbook = openpyxl.Workbook()
    site_sheet = workbook.active
    site_sheet.title = "chargers (3)"
    site_sheet.append(["name", "address", "latitude", "longitude"])

    charger_sheet = workbook.create_sheet("CT chargers")
    charger_sheet.append(
        [
            "EVSE ID", "Display Name", "MAC Address", "EVSE Latitude", "EVSE Longitude",
            "Port 1: Voltage (V)", "Port 1: Current (A)", "Port 1: Connector Type",
            "Port 2: Voltage (V)", "Port 2: Current (A)", "Port 2: Connector Type",
            "EVSE Usage Category", "Serial Number",
        ]
    )
    charger_sheet.append([1, "No Match", "mac", 0, 0, "240V", "30A", "J1772", "240V", "30A", "J1772", "cat", 123])
    workbook.save(path)

    with pytest.raises(ValueError):
        convert_workbook(str(path), default_contact_email="ops@example.com")
